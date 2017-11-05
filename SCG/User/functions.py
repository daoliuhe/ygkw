""" User-defined functions by Li Song """

import scipy.io
from openpyxl import load_workbook
import pylab as pl
import numpy as np
import itertools
import pickle
from sklearn.metrics import f1_score
from scipy.stats import pearsonr
from sklearn.preprocessing import MinMaxScaler
from scipy import signal
from sklearn.model_selection import train_test_split
from sklearn.linear_model import \
    LinearRegression, Ridge, Lasso, LassoLarsIC
from sklearn.ensemble import RandomForestRegressor
from sklearn.feature_selection import f_regression, mutual_info_regression
import pywt
import pandas as pd
from statsmodels.tsa.stattools import adfuller
from statsmodels.tsa.stattools import acf


def aicc(results):
    """
    Akaike information criterion (AIC) with small sample correction

    llf : float
        value of the loglikelihood
    nobs : int
        number of observations
    df_modelwc : int
        number of parameters including constant
    """
    llf = results.llf
    nobs = results.nobs
    df_modelwc = results.params.shape[0]
    return -2. * llf + 2. * df_modelwc * nobs / (nobs - df_modelwc - 1.)


def test_stationarity(ts):
    """
    Perform augmented Dickey-Fuller test to check stationarity
        stationary     if test statistic < critical value
        non-stationary if test statistic > critical value
    """

    print('Results of Augmented Dickey-Fuller Test:')
    test = adfuller(ts, autolag='AIC')
    out = pd.Series(test[0:4], index=['Test Statistic',
                                      'p-value', 'Lag Used',
                                      'Number of Observations Used'])
    for key, value in test[4].items():
        out['Critical Value (%s)' % key] = value
    print(out)


def test_LBQ(residuals):
    """ Perform Ljung-Box test to check residuals """

    print('Results of Ljung-Box Test:')
    acfs, q, probs = acf(residuals.values.squeeze(), qstat=True)
    data = np.c_[range(1, 41), acfs[1:], q, probs]
    table = pd.DataFrame(data, columns=['Lag', "ACF", "Q", "Prob"])
    print(table.set_index('Lag'))


def pd_sampling(df):
    """ 
    sampling to 10 min using pandas 
    15-5-2017: in Pandas, use built-in function
    """

    data_sampled = sampling(df.values[1:, :])
    start = '2016-01-01 00:10:00'
    periods = len(data_sampled)
    dates = pd.date_range(start=start, periods=periods, freq='10T')
    sampled_df2016 = pd.DataFrame(data_sampled, dates, df.columns)

    return sampled_df2016


def pd_read_csv(filename):
    """ read csv using pandas
            header: row number of variable names
            index_col: col number of data index
    """

    data = pd.read_csv(filename, header=4, index_col=1)
    start, end = data.index[0], data.index[-1]
    dates = pd.date_range(start, end, freq='T')
    df = pd.DataFrame(data.values[:, 1:], dates, data.columns[1:])

    return df


def wave_dec(x, wname, level):
    """ wavelet decomposition """

    if level == 1:
        ca, cd = pywt.dwt(x, wname)
        a = pywt.idwt(ca, None, wname)
        d = pywt.idwt(None, cd, wname)
        return a, d
    elif level == 2:
        ca2, cd2, cd1 = pywt.wavedec(x, wname, level=level)
        a2 = pywt.waverec((ca2, np.zeros(len(cd2)), np.zeros(len(cd1))), wname)
        d2 = pywt.waverec((np.zeros(len(ca2)), cd2, np.zeros(len(cd1))), wname)
        d1 = pywt.waverec((np.zeros(len(ca2)), np.zeros(len(cd2)), cd1), wname)
        return a2, d2, d1
    else:
        ca3, cd3, cd2, cd1 = pywt.wavedec(x, wname, level=level)
        a3_0 = np.zeros(len(ca3))
        d3_0 = np.zeros(len(cd3))
        d2_0 = np.zeros(len(cd2))
        d1_0 = np.zeros(len(cd1))

        a3 = pywt.waverec((ca3, d3_0, d2_0, d1_0), wname)
        d3 = pywt.waverec((a3_0, cd3, d2_0, d1_0), wname)
        d2 = pywt.waverec((a3_0, d3_0, cd2, d1_0), wname)
        d1 = pywt.waverec((a3_0, d3_0, d2_0, cd1), wname)
        return a3, d3, d2, d1


def print_time(t):
    # display time information
    if t < 60:
        print('\nRunning time: %.1f seconds' % t)
    elif 60 <= t < 3600:
        print('\nRunning time: %.f minutes %.1f seconds' % (t // 60, t % 60))
    else:
        n_hours = t // 3600
        n_mins = (t % 3600) // 60
        n_seconds = (t % 3600) % 60
        print('\nRunning time: %.f hours %.f minutes %.1f seconds'
              % (n_hours, n_mins, n_seconds))


def feature_ranking(vib, inputs, cols):
    """
    Make a ranking list of input features, using 8 scoring methods
        cols: column index of all candidate features
    """

    vib_ = MinMaxScaler().fit_transform(vib.reshape(-1, 1)).reshape(len(vib))  # (len,) to (len,1) to (len,)
    inputs_ = MinMaxScaler().fit_transform(inputs)

    ratios = [0.6, 0.7, 0.8, 0.9, 1.0]
    mean_table = np.zeros((inputs.shape[1], len(ratios)))
    for kk in range(len(ratios)):
        train_size = int(len(vib) * ratios[kk])
        vibration = vib_[20:train_size]
        inputs = inputs_[20:train_size, :]

        # 1. linear regression
        lr = LinearRegression(normalize=True)
        lr.fit(inputs, vibration)
        w1 = np.abs(lr.coef_)

        # 2. ridge regression
        ridge = Ridge(alpha=1e-4)  # regularization parameter
        ridge.fit(inputs, vibration)
        w2 = np.abs(ridge.coef_)

        # 3.4.5 lasso
        lasso1 = Lasso(alpha=1e-3)
        lasso1.fit(inputs, vibration)
        w3 = np.abs(lasso1.coef_)

        lasso2 = LassoLarsIC(criterion='aic', max_iter=50)
        lasso2.fit(inputs, vibration)
        w4 = np.abs(lasso2.coef_)

        lasso3 = LassoLarsIC(criterion='bic', max_iter=50)
        lasso3.fit(inputs, vibration)
        w5 = np.abs(lasso3.coef_)

        # 6. random forest
        rf = RandomForestRegressor(n_estimators=20)
        rf.fit(inputs, vibration)
        w6 = rf.feature_importances_

        # 7. f_regression
        f, p = f_regression(inputs, vibration)
        w7 = np.nan_to_num(f)

        # 8. mi value
        mi_values = mutual_info_regression(inputs, vibration)
        w8 = mi_values

        # mean of 8 methods
        rank_table = np.vstack((w1, w2, w3, w4, w5, w6, w7, w8)).T
        rank_scaled = MinMaxScaler().fit_transform(rank_table)
        mean_table[:, kk] = np.mean(rank_scaled, axis=1)

    mean_list = np.mean(mean_table, axis=1)
    ind = np.argsort(mean_list)[::-1]  # big to small
    ranking_cols = np.array(cols)[ind]

    return ranking_cols


def backward_selection(data, n_y_input, plot=False):
    """ Sequential backward feature selection """

    vibration = data[:, 41]
    vib = filter_data_(vibration, filter_size=6)
    train_size = int(len(vib) * 0.8)  # 80% data for training
    cols = col_index()  # index of external variables
    n = len(cols)
    removed_cols = []  # index of removed variables
    rmse_set = []
    k_set = range(1, n)  # no. of removed variables

    for k in k_set:
        remaining_cols = list(set(cols) - set(removed_cols))
        best_rmse = np.inf
        best_col = None
        for col in remaining_cols:
            augmented_cols = remaining_cols.copy()  # make a copy
            augmented_cols.remove(col)  # remove one column
            augmented = data[:, augmented_cols]  # candidate X_input
            trainX, trainY, testX, testY = prepare_data_(vib, augmented,
                                                         train_size, n_y_input)
            model = LinearRegression()
            model.fit(trainX, trainY)
            testY_pred = model.predict(testX)
            test_rmse = rmse(testY, testY_pred)

            if test_rmse < best_rmse:
                best_col = col
                best_rmse = test_rmse

        removed_cols.append(best_col)
        rmse_set.append(best_rmse)

        if k == 1:
            print('Backward feature selection progress:')
        if k % 5 == 0:
            print('%5s out of %s features have been removed' % (k, n))

    best_k = k_set[rmse_set.index(min(rmse_set))]
    best_set = list(set(cols) - set(removed_cols[:best_k]))  # remaining

    if plot:
        pl.plot(range(1, n), rmse_set[::-1])
        pl.grid()

    return n - best_k, best_set


def forward_selection(data, n_y_input, plot_rmse=False):
    """ Sequential forward feature selection """

    vibration = data[:, 41]
    vib = filter_data_(vibration, filter_size=6)
    train_size = int(len(vib) * 0.8)    # 80% data for training
    cols = col_index()                  # index of external variables
    n = len(cols)
    selected_cols = []                  # index of selected variables
    rmse_set = []
    k_set = range(1, n)

    for k in k_set:
        remaining_cols = list(set(cols) - set(selected_cols))
        best_rmse = np.inf
        best_col = None
        for col in remaining_cols:
            new_col = data[:, col].reshape(-1, 1)
            if k == 1:
                augmented = new_col
            else:
                augmented = np.hstack((X_input, new_col))
            trainX, trainY, testX, testY = prepare_data_(vib, augmented,
                                                         train_size, n_y_input)
            model = LinearRegression()
            model.fit(trainX, trainY)
            testY_pred = model.predict(testX)
            test_rmse = rmse(testY, testY_pred)

            if test_rmse < best_rmse:
                best_col = col
                best_rmse = test_rmse

        selected_cols.append(best_col)
        rmse_set.append(best_rmse)
        X_input = data[:, selected_cols]  # update X for next step

        if k == 1:
            print('Forward feature selection progress:')
        if k % 5 == 0:
            print('%5s out of %s features have been selected' % (k, n))

    best_k = k_set[rmse_set.index(min(rmse_set))]
    best_set = selected_cols[:best_k]

    if plot_rmse:
        pl.plot(range(1, n), rmse_set)
        pl.grid()

    return best_k, best_set


def col_index():
    """
    Alternative to return index for the variables, because we know
    the index number of variables to be removed
        variable        index
        TI392-03A.PV    3
        PDI382-05.PV    34
        XI392-04A.PV    41
        XI392-04B.PV    42

    Original: 66 (from excel file)
    Remove repeated: PDI382-05.PV, TI392-03A.PV
    Now: 66 - 2 = 64
    Remove vibrations: XI392-04A.PV, XI392-04B.PV
    Now: 64 - 2 = 62
    """

    # headers.p is dict, key: column index, value: variable name
    headers = pickle.load(open("../Data/headers.p", "rb"))
    del headers[3], headers[34], headers[41], headers[42]
    return list(headers.keys())


def select_lag(vib, train_size, horizon):
    """ Select best number of lagged vibration inputs """

    best_lag = None
    best_rmse = np.inf
    max_lag = 50
    lags = list(range(horizon, max_lag + 1))
    for lag in lags:
        X, Y, _, __ = prepare_data(vib, train_size, lag)
        trainX, valX, trainY, valY = train_test_split(X, Y, test_size=0.3, random_state=42)
        model = LinearRegression()
        model.fit(trainX, trainY)
        valY_pred = model.predict(valX)
        val_rmse = rmse(valY, valY_pred)

        if val_rmse < best_rmse:
            best_lag = lag
            best_rmse = val_rmse

    return best_lag


def persistence(year='2016', start='14/Feb/16 00:10:00', end='25/Mar/16 00:10:00'):
    """ Persistence method: benchmark model """

    if year == '2014':
        data_sampled = pickle.load(open("../Data/sampled2014.p", "rb"))
        date_sampled = pickle.load(open("../Data/date2014.p", "rb"))
        data = slice_data(date_sampled, data_sampled, start, end)
    else:
        data_sampled = pickle.load(open("../Data/sampled2016.p", "rb"))
        date_sampled = pickle.load(open("../Data/date2016.p", "rb"))
        data = slice_data(date_sampled, data_sampled, start, end)

    vibration = data[:, 41]
    train_size = 3600
    testY = vibration[train_size:]
    testY_pred = np.zeros(len(testY))
    horizon = 6
    n_hours = int(len(testY) / horizon)  # for hourly update
    y = vibration[train_size-1:]
    for hour in range(n_hours):
        start = hour * horizon
        testY_pred[start:start + horizon] = np.ones(horizon) * y[start]
    test_rmse = rmse(testY, testY_pred)

    return test_rmse


def mape(y_true, y_pred):
    return np.mean(abs((y_true - y_pred) / y_true)) * 100


def filter_data_(vib, filter_size=5):
    """ Apply forward-backward linear filter """

    b, a = (1 / filter_size) * np.ones(filter_size), 1
    return signal.filtfilt(b, a, vib)

# -----------------------------------------
# 3/1/2017: copy Dec/func to here, it's better to put all functions together


def prepare_data_(y, X, train_size, n_y_input):
    """ Prepare training and testing data, including external variables X """

    y_lags = np.array(range(n_y_input)) + 1  # list not work, use array
    n_x_input = X.shape[1]
    n_input = n_y_input + n_x_input

    trainY = y[n_y_input:train_size]
    testY = y[train_size:]
    trainX = np.zeros((len(trainY), n_input))
    testX = np.zeros((len(testY), n_input))

    h1 = range(n_y_input, train_size)
    h2 = range(train_size, len(y))
    for i in range(0, n_input):
        if i < n_y_input:
            trainX[:, i] = y[h1 - y_lags[i]]
            testX[:, i] = y[h2 - y_lags[i]]
        else:
            trainX[:, i] = X[:, i - n_y_input][h1 - y_lags[0]]  # only lag 1 is used
            testX[:, i] = X[:, i - n_y_input][h2 - y_lags[0]]

    return trainX, trainY, testX, testY


def iterative(model, X, H, n_y_input):
    """
        model: forecasting model
        X: input batch, including external features
        H: horizon

    13/1/2017: as direct forecasting is used always, this function is no longer used.
    """

    forecast = np.zeros(H)
    forecast[0] = model.predict(X[0, :].reshape(1, -1))  # 1-step ahead forecast

    for h in range(1, H):
        features = np.hstack((forecast[:h], X[h, :(n_y_input-h)], X[h, n_y_input:]))
        forecast[h] = model.predict(features.reshape(1, -1))
    return forecast


def print_progress(step, max_step, stride=10):
    if step % stride == 0:
        # print progress every 10 steps
        print('Progress: %d/%d' % (step, max_step))


def mimo_data(series, window, H):
    """ data preparation for mimo model """

    n = len(series)
    n_samples = n - H - window + 1
    inputs = np.zeros((n_samples, window))
    targets = np.zeros((n_samples, H))

    for t in range(n_samples):
        inputs[t, :] = series[t:(t + window)]
        targets[t, :] = series[(t + window):(t + window + H)]

    return inputs, targets


def iterative_forecast(model, X, H, window):
    """
        model: forecasting model
        X: input batch, excluding external variables
        H: horizon
        window: no. of features
    """

    forecast = np.zeros(H)
    forecast[0] = model.predict(X[0, :].reshape(1, -1))  # 1-step ahead forecast

    for h in range(1, H):
        augmented = np.hstack((forecast[:h], X[h, :]))
        features = augmented[:window]
        forecast[h] = model.predict(features.reshape(1, -1))

    return forecast


def direct_forecast(model, trainX, trainY, testX, H):
    """
        model: forecasting model
        trainX, testX: input batch, can include or exclude external variables
        H: horizon
    """

    forecast = np.zeros(H)
    n = len(trainY)
    n_samples = n - H + 1

    for h in range(H):
        X = trainX[0:n_samples, :]
        y = trainY[(0+h):(n_samples+h)]
        model.fit(X, y)
        forecast[h] = model.predict(testX.reshape(1, -1))

    return forecast


def prepare_data(vib, train_size, n_y_input):
    """ Prepare training and testing data, excluding external variables """

    y_lags = np.array(range(n_y_input)) + 1  # list will not work, type is array
    trainY = vib[n_y_input:train_size]
    testY = vib[train_size:]
    trainX = np.zeros((len(trainY), n_y_input))
    testX = np.zeros((len(testY), n_y_input))

    h1 = range(n_y_input, train_size)
    h2 = range(train_size, len(vib))
    for i in range(n_y_input):
        trainX[:, i] = vib[h1 - y_lags[i]]
        testX[:, i] = vib[h2 - y_lags[i]]

    return trainX, trainY, testX, testY


def remove_low_r_var(y, X):
    """ Remove features with low r and var """

    col_ind = pickle.load(open("../Nov/Regression/col_index_new.p", "rb"))

    # remove features when r < 0.1
    r_limit = 0.1
    r = pcc(X, y)
    r_ind = r >= r_limit
    inputs1 = X[:, r_ind]
    col_ind1 = np.array(col_ind)[r_ind]

    # remove features when var < 0.005, do scaling first
    inputs2 = MinMaxScaler().fit_transform(inputs1)
    var = variance(inputs2)
    var_limit = 0.005
    var_ind = var >= var_limit
    col_ind2 = col_ind1[var_ind]

    return list(col_ind2)


def filter_data(filter_size=5, start='14/Feb/16 00:10:00', end='25/Mar/16 00:10:00'):
    """ Apply forward-backward linear filter """

    data_sampled = pickle.load(open("../Data/sampled2016.p", "rb"))
    date_sampled = pickle.load(open("../Data/date2016.p", "rb"))
    data = slice_data(date_sampled, data_sampled, start, end)

    vib = data[:, 41]  # XI392-04A
    b = (1 / filter_size) * np.ones(filter_size)
    a = 1
    y = signal.filtfilt(b, a, vib)

    col_ind = pickle.load(open("../Nov/Regression/col_index_new.p", "rb"))
    inputs = data[:, col_ind]
    X = signal.filtfilt(b, a, inputs, axis=0)

    return y, X
# ----------------------------------------------------


def aic(y_true, y_pred, n_samples, n_features):
    mse = np.mean((y_true - y_pred) ** 2)
    aic = n_samples * np.log(mse) + 2 * n_features

    return aic


def rmse(y_true, y_pred):
    return np.sqrt(np.mean((y_true - y_pred) ** 2))


def top_k(x, k):
    """ return big k elements in a vector"""

    ind = np.argsort(x)[::-1][:k]
    return x[ind]


def remove_low_var_pcc(r_limit=0.1, var_limit=0.005):
    """ Remove features with low variances and pearson r """

    data_sampled = pickle.load(open("../Data/sampled2016.p", "rb"))
    date_sampled = pickle.load(open("../Data/date2016.p", "rb"))
    date_start, date_end = '01/Feb/16 00:10:00', '01/Jun/16 00:10:00'
    data = slice_data(date_sampled, data_sampled, date_start, date_end)
    vib = data[:, 41]  # XI392-04A

    col_ind = pickle.load(open("../Nov/Regression/col_index_new.p", "rb"))
    inputs = data[:, col_ind]

    # remove features when r < 0.1
    r = pcc(inputs, vib)
    r_ind = r >= r_limit
    inputs1 = inputs[:, r_ind]
    col_ind1 = np.array(col_ind)[r_ind]

    # remove features when var < 0.005
    # do scaling first
    sx = MinMaxScaler()
    inputs2 = sx.fit_transform(inputs1)
    var = variance(inputs2)
    var_ind = var >= var_limit
    inputs3 = inputs1[:, var_ind]
    col_ind2 = col_ind1[var_ind]

    return list(col_ind2)


def variance(X):
    """ Compute variances of columns of X """
    var = [np.var(x) for x in X.T]
    return np.array(var)


def read_mat(filename):
    # read data from mat file
    data = scipy.io.loadmat(filename)
    data = data['data']

    return data


def read_header(filename, start_row=5, start_col=3):
    """
    read variable names from excel

    filename: excel file
    start_row: starting row of variable names, same as excel row no.
    start_col: starting col of variable names, same as excel col no.

    return: variable names, dict
    """

    wb = load_workbook(filename, read_only=True)
    sheet = wb['Sheet1']

    headers = {}
    for col in range(start_col, sheet.max_column+1):
        headers[col - start_col] = sheet.cell(row=start_row, column=col).value

    return headers


def read_datetime(filename):
    """
    one datetime <--> one row number

    row number for indexing

    6/11/2016
    use to produce date2014.p, date2016.p
    """

    data = [line.rstrip('\n') for line in open(filename)]
    data_sampled = data[9::10]  # sampling

    date = {}
    for row in range(0, len(data_sampled)):
        date[data_sampled[row]] = row

    return date


def sampling(data, interval=10):
    """
    Perform down sampling for data

    01/Mar/14 00:01:00 \
    01/Mar/14 00:02:00  \
    01/Mar/14 00:03:00   \
    01/Mar/14 00:04:00    \
    01/Mar/14 00:05:00     \
    01/Mar/14 00:06:00       -->>> 01/Mar/14 00:10:00
    01/Mar/14 00:07:00    /
    01/Mar/14 00:08:00   /
    01/Mar/14 00:09:00  /
    01/Mar/14 00:10:00 /


    data: original data
    interval: sampling rate
    """

    data_period = data[1:, :]  # skip 1st row

    m = round(data_period.shape[0] / 10)
    n = data_period.shape[1]
    data_sampled = np.zeros((m, n))

    for j in range(0, n):
        for i in range(0, m):
            data_sampled[i, j] = np.mean(data_period[10 * i:10 * (i + 1), j])

    return data_sampled


class Data(object):
    """ A simulated structure, consisting datetime and data """

    def __init__(self, date, data):
        self.date = date
        self.data = data


def slice_data(date_sampled, data_sampled, date_start, date_end):

    data = Data(date_sampled, data_sampled)
    row_start = data.date[date_start]
    row_end = data.date[date_end]
    data_sliced = data.data[row_start:row_end, :]

    return data_sliced


def plot_confusion_matrix(con_matrix, classes):

    """
    plot confusion matrix to evaluate the accuracy of classification

    For a confusion matrix C, the element C_ij is equal to the number of observations
    known to be in label i but predicted to be in label j.
    """

    pl.imshow(con_matrix, interpolation='nearest', cmap=pl.cm.Blues)
    # pl.title('Confusion Matrix')
    pl.colorbar()
    tick_marks = np.arange(len(classes))

    thresh = con_matrix.max() / 2.
    for i, j in itertools.product(range(con_matrix.shape[0]), range(con_matrix.shape[1])):
        pl.text(j, i, con_matrix[i, j], fontsize=15, horizontalalignment="center",
                color="white" if con_matrix[i, j] > thresh else "black")

    pl.xticks(tick_marks, classes, fontsize=15)
    pl.yticks(tick_marks, classes, fontsize=15)
    pl.ylabel('True label', fontsize=15)
    pl.xlabel('Predicted label', fontsize=15)


def read_data(year='2016', start='14/Feb/16 00:10:00', end='25/Mar/16 00:10:00'):
    """
    29/11/2016: read data function, save lines
    year: 2014 or 2016
    return: y (vibration), X (features)

    29/12/2016
    Major changes
    return: data (vibration + features)
    """

    if year == '2014':
        data_sampled = pickle.load(open("../Data/sampled2014.p", "rb"))
        date_sampled = pickle.load(open("../Data/date2014.p", "rb"))
    else:
        data_sampled = pickle.load(open("../Data/sampled2016.p", "rb"))
        date_sampled = pickle.load(open("../Data/date2016.p", "rb"))

    data = slice_data(date_sampled, data_sampled, start, end)

    return data


def pdf(x, mean, std):
    """
    Compute probability given x
    30/11/2016: double checked, formula below is correct.
    """

    return np.exp(-0.5*((x-mean)/std)**2)/np.sqrt(2*np.pi*std*std)


def pick_epsilon(pval, valY):
    """
    select best threshold

    pval: probabilities of validation data
    valY: true labels of validation data
    return: best epsilon
    """

    best_epsilon = 0
    best_f1 = 0
    f1 = 0
    step = 10000
    for epsilon in np.linspace(min(pval), max(pval), step):
        valY_pred = pval < epsilon
        f1 = f1_score(valY, valY_pred)

        if f1 > best_f1:
            best_epsilon = epsilon
            best_f1 = f1

    print('Best epsilon: %s' % best_epsilon)
    print('Best f1 score: %s' % best_f1)
    return best_epsilon


def compute_prob(X, mu, std):
    """ compute probabilities of given samples """

    # new implementation: for loop in one line
    p = [np.prod(pdf(x, mu, std)) for x in X]

    return np.array(p)


def find_name(cols):
    """ Given col number, find variable name """

    headers = pickle.load(open("../Data/headers.p", "rb"))
    for col in cols:
        print(headers[col])


def pcc(X, y):
    """
    PCC: Pearson correlation coefficient

    X: input matrix
    y: target vector
    return: Pearson r
    """

    # for loop in one line, very neat
    r = [np.abs(pearsonr(x, y)[0]) for x in X.T]  # column by column
    return np.array(r)

